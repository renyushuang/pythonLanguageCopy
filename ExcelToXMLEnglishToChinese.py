import openpyxl, os
from xml.dom import minidom

'''
需要安装 openpyxl 即命令 sudo pip install openpyxl
参数 
--pathExcel excel的绝对路径  
--pathEnglishStrings英文文案的文件夹路径 
--fileName 英文文案的文件名
示例
python3 ExcelToXMLEnglishToChinese.py --pathExcel=./work/VideoDownloaderV1.0.0.xlsx --pathEnglishStrings=./work --fileName=strings.xml
'''

pathExcel = ''
pathEnglishStrings = ''
fileName = ''


def __get_params():
    from optparse import OptionParser
    parser = OptionParser(description="Add new translation")
    parser.add_option("--pathExcel", help="pathExcel", metavar="FILE")
    parser.add_option("--pathEnglishStrings", help="pathEnglishStrings")
    parser.add_option("--fileName", help="string file name")
    (options, args) = parser.parse_args()
    print(options)
    if options.pathExcel is None:
        print("no pathExcel file")
        exit(-1)

    if options.pathEnglishStrings is None:
        print("no pathEnglishStrings Dir")
        exit(-1)

    if options.fileName is None:
        print("no file name")
        exit(-1)
    global pathExcel
    global pathEnglishStrings
    global fileName
    pathExcel = options.pathExcel
    pathEnglishStrings = options.pathEnglishStrings
    fileName = options.fileName


def getExcelChineseAndEnglishMap():
    map = {}
    # 读取Excel
    wb = openpyxl.load_workbook(pathExcel)
    sheetNames = wb.sheetnames  # 获取工作表
    for sheetName in sheetNames:
        sheet = wb[sheetName]
        for row in range(1, sheet.max_row):
            map['CD' + str(row)] = {'chinese': sheet['C' + str(row)].value, 'english': sheet['D' + str(row)].value, }
    return map


def writeXmlNewString(map):
    doc = minidom.Document()

    root = minidom.parse(os.path.join(pathEnglishStrings, fileName))

    docRoot = doc.createElement('resources')
    docRoot.setAttribute('xmlns:tools', 'http://schemas.android.com/tools')
    docRoot.setAttribute('xmlns:xliff', 'urn:oasis:names:tc:xliff:document:1.2')
    doc.appendChild(docRoot)
    # 将原有的内容copy一遍
    elements = root.getElementsByTagName('string')
    for ele in elements:
        element = doc.createElement('string')
        element.setAttribute("name", ele.attributes['name'].value)
        data = None
        # 将新增加的内容copy进来
        for k, value in map.items():
            english = value['english']
            if english.strip() == ele.firstChild.data.strip():
                data = doc.createTextNode(value['chinese'])
                element.appendChild(data)

        if data == None:
            element.appendChild(doc.createTextNode(ele.firstChild.data))

        doc.documentElement.appendChild(element)

    # 新建的xml进行copy
    fp = open(os.path.join(pathEnglishStrings, 'strings_ch.xml'), 'w')
    doc.writexml(fp, indent='', addindent='\t', newl='\n')


if __name__ == '__main__':
    __get_params()
    writeXmlNewString(getExcelChineseAndEnglishMap())
